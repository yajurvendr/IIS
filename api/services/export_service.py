"""Excel and PDF export builders."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PRIMARY_FILL = PatternFill(fill_type="solid", fgColor="1A3C5E")
WHITE_FONT   = Font(bold=True, color="FFFFFF")
CENTER       = Alignment(horizontal="center", vertical="center")

RED_FILL   = PatternFill(fill_type="solid", fgColor="FF0000")
AMBER_FILL = PatternFill(fill_type="solid", fgColor="FFA500")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="00AA00")


def _apply_header(ws, headers: list[tuple]):
    """headers: [(label, width), ...]"""
    ws.append([h[0] for h in headers])
    for i, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.font      = WHITE_FONT
        cell.fill      = PRIMARY_FILL
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 20


def _woi_fill(status: str):
    if status == "red":   return RED_FILL
    if status == "amber": return AMBER_FILL
    return GREEN_FILL


def build_po_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Recommendation"

    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Unit", 8), ("Current Stock", 14), ("DRR (rec)", 12), ("WOI (wks)", 10),
        ("WOI Status", 12), ("MSL Suggested", 14), ("Target 12W Qty", 14),
        ("Order Qty", 12), ("Purchase Cost", 14), ("Est. PO Value", 14),
    ]
    _apply_header(ws, headers)

    for r in rows:
        po_val = (r.get("suggested_order_qty") or 0) * (r.get("purchase_cost_decoded") or 0)
        data_row = ws.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            r.get("unit"), r.get("current_stock") or 0,
            round(float(r.get("drr_recommended") or 0), 2),
            round(float(r.get("woi") or 0), 1),
            (r.get("woi_status") or "").upper(),
            r.get("msl_suggested") or 0, r.get("target_12w_qty") or 0,
            r.get("suggested_order_qty") or 0,
            round(float(r.get("purchase_cost_decoded") or 0), 2),
            round(po_val, 2),
        ])
        row_idx = ws.max_row
        status  = (r.get("woi_status") or "").lower()
        cell    = ws.cell(row=row_idx, column=9)
        cell.fill = _woi_fill(status)
        cell.font = Font(bold=True, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_inventory_woi_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory WOI"

    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Unit", 8), ("Current Stock", 14), ("DRR (rec)", 12), ("WOI (wks)", 10),
        ("WOI Status", 12), ("MSL Suggested", 14), ("Order Qty", 12), ("Forecast At", 18),
    ]
    _apply_header(ws, headers)

    for r in rows:
        ws.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            r.get("unit"), r.get("current_stock") or 0,
            round(float(r.get("drr_recommended") or 0), 2),
            round(float(r.get("woi") or 0), 1),
            (r.get("woi_status") or "").upper(),
            r.get("msl_suggested") or 0,
            r.get("suggested_order_qty") or 0,
            str(r.get("computed_at") or "")[:19],
        ])
        row_idx = ws.max_row
        status  = (r.get("woi_status") or "").lower()
        cell    = ws.cell(row=row_idx, column=9)
        cell.fill = _woi_fill(status)
        cell.font = Font(bold=True, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_msl_review_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MSL Review"

    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Busy MSL", 12), ("System MSL", 12), ("Variance", 12),
        ("Current Stock", 14), ("DRR (rec)", 12), ("WOI Status", 12), ("Recommendation", 18),
    ]
    _apply_header(ws, headers)

    LIGHT_GREEN = PatternFill(fill_type="solid", fgColor="CCFFCC")
    LIGHT_RED   = PatternFill(fill_type="solid", fgColor="FFCCCC")

    for r in rows:
        busy   = float(r.get("busy_msl") or 0)
        system = float(r.get("system_msl") or 0)
        var    = float(r.get("variance") or 0)
        if busy > 0 and system > busy * 1.2:
            rec = "Increase MSL"
        elif busy > 0 and system < busy * 0.8:
            rec = "Reduce MSL"
        else:
            rec = "OK"

        ws.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            busy, system, round(var, 2),
            r.get("current_stock") or 0,
            round(float(r.get("drr_recommended") or 0), 2),
            (r.get("woi_status") or "").upper(),
            rec,
        ])
        row_idx = ws.max_row
        if rec == "Increase MSL":
            for col in range(5, 12):
                ws.cell(row=row_idx, column=col).fill = LIGHT_RED
        elif rec == "Reduce MSL":
            for col in range(5, 12):
                ws.cell(row=row_idx, column=col).fill = LIGHT_GREEN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_profitability_excel(summary, by_category, by_brand, top_skus) -> bytes:
    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _apply_header(ws, [("Metric", 28), ("Value", 20)])
    s = summary or {}
    for label, key in [
        ("Total Revenue", "total_revenue"),
        ("Total COGS", "total_cogs"),
        ("Gross Profit", "gross_profit"),
        ("Total Quantity Sold", "total_qty"),
    ]:
        v = float(s.get(key) or 0)
        ws.append([label, round(v, 2)])

    # ── By Category ──────────────────────────────────────────────────────────
    ws_cat = wb.create_sheet("By Category")
    _apply_header(ws_cat, [("Category", 24), ("Revenue", 18), ("COGS", 18), ("Gross Profit", 18), ("Margin %", 12)])
    for r in (by_category or []):
        rev  = float(r.get("revenue") or 0)
        gp   = float(r.get("gross_profit") or 0)
        pct  = round(gp / rev * 100, 1) if rev > 0 else 0
        ws_cat.append([r.get("category"), round(rev, 2), round(float(r.get("cogs") or 0), 2), round(gp, 2), pct])

    # ── By Brand ─────────────────────────────────────────────────────────────
    ws_brand = wb.create_sheet("By Brand")
    _apply_header(ws_brand, [("Brand", 24), ("Revenue", 18), ("COGS", 18), ("Gross Profit", 18), ("Margin %", 12)])
    for r in (by_brand or []):
        rev  = float(r.get("revenue") or 0)
        gp   = float(r.get("gross_profit") or 0)
        pct  = round(gp / rev * 100, 1) if rev > 0 else 0
        ws_brand.append([r.get("brand"), round(rev, 2), round(float(r.get("cogs") or 0), 2), round(gp, 2), pct])

    # ── Top SKUs ─────────────────────────────────────────────────────────────
    ws_sku = wb.create_sheet("Top SKUs")
    _apply_header(ws_sku, [("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
                            ("Revenue", 16), ("COGS", 16), ("Gross Profit", 16), ("Margin %", 12), ("Qty", 12)])
    for r in (top_skus or []):
        rev = float(r.get("revenue") or 0)
        gp  = float(r.get("gross_profit") or 0)
        pct = round(gp / rev * 100, 1) if rev > 0 else 0
        ws_sku.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            round(rev, 2), round(float(r.get("cogs") or 0), 2), round(gp, 2), pct, r.get("qty") or 0,
        ])

    ws_sku.freeze_panes = "A2"
    ws_sku.auto_filter.ref = f"A1:I{ws_sku.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pre_season_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-Season Alert"
    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Current Stock", 14), ("DRR (rec)", 12), ("WOI (wks)", 10), ("WOI Status", 12),
        ("Suggested Order Qty", 18), ("Latest Order Date", 18),
    ]
    _apply_header(ws, headers)
    ORANGE_FILL = PatternFill(fill_type="solid", fgColor="FF6B00")
    for r in rows:
        ws.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            r.get("current_stock") or 0,
            round(float(r.get("drr_recommended") or 0), 2),
            round(float(r.get("woi") or 0), 1),
            (r.get("woi_status") or "").upper(),
            r.get("suggested_order_qty") or 0,
            str(r.get("latest_order_date") or "")[:10],
        ])
        row_idx = ws.max_row
        cell = ws.cell(row=row_idx, column=8)
        cell.fill = _woi_fill(r.get("woi_status") or "")
        cell.font = Font(bold=True, color="FFFFFF")
        # Highlight if latest order date is within 14 days
        lod = str(r.get("latest_order_date") or "")
        if lod:
            ws.cell(row=row_idx, column=10).fill = ORANGE_FILL
            ws.cell(row=row_idx, column=10).font = Font(bold=True, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_volume_profit_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volume-Profit Divergence"
    headers = [
        ("Vol Rank", 10), ("SKU Code", 16), ("SKU Name", 32),
        ("Brand", 16), ("Category", 16),
        ("Total Qty", 14), ("Revenue", 16), ("Margin %", 12),
    ]
    _apply_header(ws, headers)
    WARN_FILL = PatternFill(fill_type="solid", fgColor="FFD700")
    for r in rows:
        pct = round(float(r.get("margin_pct") or 0), 1)
        ws.append([
            r.get("vol_rank"), r.get("sku_code"), r.get("sku_name"),
            r.get("brand"), r.get("category"),
            round(float(r.get("total_qty") or 0), 2),
            round(float(r.get("revenue") or 0), 2),
            pct,
        ])
        row_idx = ws.max_row
        # Highlight rows with margin < 10%
        if pct < 10:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = WARN_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sales_forecast_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Unit", 8), ("Current Stock", 14), ("DRR (rec)", 12), ("WOI (wks)", 10),
        ("WOI Status", 12), ("Proj. 4W Stock", 14), ("Proj. 8W Stock", 14),
        ("Proj. 12W Stock", 15), ("Stock-Out Date", 16), ("Suggested Order", 15),
    ]
    _apply_header(ws, headers)

    STOCKOUT_FILL = PatternFill(fill_type="solid", fgColor="FFE0E0")
    for r in rows:
        ws.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            r.get("unit"),
            round(float(r.get("current_stock") or 0), 0),
            round(float(r.get("drr_recommended") or 0), 2),
            round(float(r.get("woi") or 0), 1),
            (r.get("woi_status") or "").upper(),
            round(float(r.get("proj_4w") or 0), 0),
            round(float(r.get("proj_8w") or 0), 0),
            round(float(r.get("proj_12w") or 0), 0),
            str(r.get("stockout_date") or "")[:10],
            r.get("suggested_order_qty") or 0,
        ])
        row_idx = ws.max_row
        status = (r.get("woi_status") or "").lower()
        cell = ws.cell(row=row_idx, column=9)
        cell.fill = _woi_fill(status)
        cell.font = Font(bold=True, color="FFFFFF")
        # Highlight projected zero stock cells
        for col, key in [(10, "proj_4w"), (11, "proj_8w"), (12, "proj_12w")]:
            if float(r.get(key) or 0) == 0:
                ws.cell(row=row_idx, column=col).fill = STOCKOUT_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_top300_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top 300 by Sales"

    headers = [
        ("Rank", 8), ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Total Qty", 12), ("Revenue", 16), ("Margin %", 12),
        ("Current Stock", 14), ("WOI Status", 12),
    ]
    _apply_header(ws, headers)

    LOW_MARGIN = PatternFill(fill_type="solid", fgColor="FFF3CD")
    for r in rows:
        pct = round(float(r.get("margin_pct") or 0), 1)
        ws.append([
            r.get("rank"),
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            round(float(r.get("total_qty") or 0), 0),
            round(float(r.get("revenue") or 0), 2),
            pct,
            round(float(r.get("current_stock") or 0), 0),
            (r.get("woi_status") or "").upper(),
        ])
        row_idx = ws.max_row
        status = (r.get("woi_status") or "").lower()
        cell = ws.cell(row=row_idx, column=10)
        cell.fill = _woi_fill(status)
        cell.font = Font(bold=True, color="FFFFFF")
        if pct < 10:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = LOW_MARGIN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_focus_sku_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Focus SKU Report"

    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16), ("Category", 16),
        ("Unit", 8), ("MSL", 12), ("Current Stock", 14),
        ("DRR (rec)", 12), ("WOI (wks)", 10), ("WOI Status", 12),
        ("System MSL", 12), ("Suggested Order", 15),
    ]
    _apply_header(ws, headers)

    for r in rows:
        ws.append([
            r.get("sku_code"), r.get("sku_name"), r.get("brand"), r.get("category"),
            r.get("unit"),
            round(float(r.get("msl") or 0), 0),
            round(float(r.get("current_stock") or 0), 0),
            round(float(r.get("drr_recommended") or 0), 2),
            round(float(r.get("woi") or 0), 1),
            (r.get("woi_status") or "").upper(),
            round(float(r.get("msl_suggested") or 0), 0),
            r.get("suggested_order_qty") or 0,
        ])
        row_idx = ws.max_row
        status = (r.get("woi_status") or "").lower()
        cell = ws.cell(row=row_idx, column=10)
        cell.fill = _woi_fill(status)
        cell.font = Font(bold=True, color="FFFFFF")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_transfer_log_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Transfer Log"
    headers = [
        ("Date", 14), ("SKU Code", 16), ("SKU Name", 32),
        ("From Branch", 20), ("To Branch", 20),
        ("Qty", 10), ("Transferred By", 22), ("Notes", 30),
    ]
    _apply_header(ws, headers)
    for r in rows:
        ws.append([
            str(r.get("transfer_date") or "")[:10],
            r.get("sku_code"), r.get("sku_name"),
            r.get("from_branch"), r.get("to_branch"),
            round(float(r.get("quantity") or 0), 2),
            r.get("transferred_by") or "",
            r.get("notes") or "",
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_outstanding_pdf(rows: list[dict], ageing: dict | None = None) -> bytes:
    """Generate an A4 PDF for the Customer Outstanding report."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    # ── Header ────────────────────────────────────────────────────────────────
    pdf.set_fill_color(26, 60, 94)
    pdf.rect(0, 0, 297, 18, "F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_xy(8, 4)
    pdf.cell(0, 8, "Customer Outstanding Report", ln=0)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_xy(220, 4)
    from datetime import date
    pdf.cell(0, 8, f"Generated: {date.today().strftime('%d %b %Y')}", ln=0)
    pdf.set_text_color(0, 0, 0)

    # ── Ageing Summary ────────────────────────────────────────────────────────
    if ageing:
        pdf.set_xy(8, 22)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(240, 244, 255)
        buckets = [
            ("0-30 Days", ageing.get("bucket_0_30", 0)),
            ("31-60 Days", ageing.get("bucket_31_60", 0)),
            ("61-90 Days", ageing.get("bucket_61_90", 0)),
            ("90+ Days",   ageing.get("bucket_90plus", 0)),
        ]
        for label, val in buckets:
            pdf.set_fill_color(240, 244, 255)
            pdf.cell(60, 8, f"{label}: ₹{float(val):,.0f}", border=1, fill=True, ln=0)
        pdf.ln(10)

    # ── Table Header ──────────────────────────────────────────────────────────
    y = pdf.get_y() + 4
    pdf.set_xy(8, y)
    pdf.set_fill_color(26, 60, 94)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    cols = [
        ("Customer Name", 60), ("Code", 22), ("Phone", 30),
        ("Total Outstanding", 36), ("Invoices", 20), ("Max Overdue (days)", 38),
        ("Ageing Bucket", 35),
    ]
    for label, w in cols:
        pdf.cell(w, 7, label, border=0, fill=True, ln=0)
    pdf.ln(7)
    pdf.set_text_color(0, 0, 0)

    # ── Rows ──────────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "", 8)
    for i, r in enumerate(rows):
        overdue = int(r.get("max_overdue_days") or 0)
        bucket = ("0-30 days" if overdue <= 30 else
                  "31-60 days" if overdue <= 60 else
                  "61-90 days" if overdue <= 90 else "90+ days")
        outstanding = float(r.get("total_outstanding") or 0)

        if overdue > 90:
            pdf.set_fill_color(255, 204, 204)
        elif overdue > 60:
            pdf.set_fill_color(255, 244, 204)
        else:
            pdf.set_fill_color(255, 255, 255) if i % 2 == 0 else pdf.set_fill_color(248, 250, 252)

        fill = True
        pdf.cell(60, 6, str(r.get("name") or "")[:35], border=0, fill=fill, ln=0)
        pdf.cell(22, 6, str(r.get("customer_code") or "")[:10], border=0, fill=fill, ln=0)
        pdf.cell(30, 6, str(r.get("phone") or "")[:15], border=0, fill=fill, ln=0)
        pdf.cell(36, 6, f"₹{outstanding:,.0f}", border=0, fill=fill, ln=0)
        pdf.cell(20, 6, str(r.get("invoice_count") or 0), border=0, fill=fill, ln=0)
        pdf.cell(38, 6, str(overdue), border=0, fill=fill, ln=0)
        pdf.cell(35, 6, bucket, border=0, fill=fill, ln=0)
        pdf.ln(6)

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


def build_outstanding_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding Report"

    headers = [
        ("Customer Name", 28), ("Customer Code", 16), ("Phone", 14),
        ("Invoice No", 16), ("Invoice Amount", 16), ("Balance Due", 14),
        ("Due Date", 14), ("Overdue Days", 14), ("Ageing Bucket", 16),
    ]
    _apply_header(ws, headers)

    LIGHT_RED  = PatternFill(fill_type="solid", fgColor="FFCCCC")
    LIGHT_AMB  = PatternFill(fill_type="solid", fgColor="FFF4CC")

    for r in rows:
        overdue = int(r.get("overdue_days") or 0)
        bucket  = ("0-30 days" if overdue <= 30 else
                   "31-60 days" if overdue <= 60 else
                   "61-90 days" if overdue <= 90 else "90+ days")
        due_str = str(r.get("due_date") or "")[:10]
        ws.append([
            r.get("customer_name"), r.get("customer_code"), r.get("phone"),
            r.get("invoice_no"), float(r.get("amount") or 0), float(r.get("balance") or 0),
            due_str, overdue, bucket,
        ])
        if overdue > 60:
            row_idx = ws.max_row
            fill = LIGHT_RED if overdue > 90 else LIGHT_AMB
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = fill

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
