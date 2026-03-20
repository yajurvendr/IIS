"""Reorder module — Smart Reorder Screen (SRS Note v2 §2 / SRS §7B).

Two analysis buckets shown together on /reorder:
  Bucket 1 — Priority SKUs: sold in last 7 days + msl > 0
  Bucket 2 — Broad Scan  : all other at-risk SKUs with msl > 0, NOT in Bucket 1

Suggested reorder qty formula (confirmed):
  lead_time_demand  = DRR_recommended × lead_time_days
  safety_buffer     = lead_time_demand × 0.20
  target_stock      = lead_time_demand + safety_buffer
  suggested_order   = MAX(0, target_stock − effective_stock)
"""
from __future__ import annotations
import io
import uuid
from datetime import date, timedelta
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config.db import fetchall, fetchone, execute, get_public_pool
from middleware.auth import require_role, get_tenant_db

router = APIRouter(prefix="/reorder", tags=["reorder"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _suggested_qty(drr_recommended, lead_time_days, effective_stock) -> int:
    drr  = float(drr_recommended or 0)
    ltd  = int(lead_time_days or 0)
    stk  = float(effective_stock or 0)
    lead_demand  = drr * ltd
    target_stock = lead_demand * 1.20
    return max(0, round(target_stock - stk))


def _expected_delivery(lead_time_days: int) -> date:
    return date.today() + timedelta(days=lead_time_days)


# ── Base reorder query fragments ──────────────────────────────────────────────

_REORDER_SELECT = """
    SELECT
        s.id AS sku_id, s.sku_code, s.sku_name, s.brand, s.category, s.unit,
        sm.msl,
        fc.drr_recommended, fc.woi, fc.woi_status, fc.msl_suggested,
        fc.current_stock AS effective_stock,
        fc.computed_at,
        -- Latest open order for this sku+branch+godown (if any)
        ro.id          AS order_id,
        ro.ordered_qty,
        ro.order_placed_at,
        ro.expected_delivery_dt,
        ro.status      AS order_status,
        ro.notes       AS order_notes
"""

_REORDER_FROM = """
    FROM skus s
    JOIN sku_msl sm ON sm.sku_id = s.id AND sm.msl > 0
    LEFT JOIN forecasting_cache fc ON fc.sku_id = s.id
    LEFT JOIN LATERAL (
        SELECT id, ordered_qty, order_placed_at, expected_delivery_dt, status, notes
        FROM skus_reorder_orders sro
        WHERE sro.sku_id = s.id
          AND sro.branch_id = sm.branch_id
          AND (sro.godown_id = sm.godown_id OR (sro.godown_id IS NULL AND sm.godown_id IS NULL))
          AND sro.status NOT IN ('delivered','cancelled')
        ORDER BY sro.created_at DESC
        LIMIT 1
    ) ro ON TRUE
"""


def _branch_godown_cond(branch_id: str, godown_id: str) -> tuple:
    conds, params = [], []
    if branch_id:
        conds.append("sm.branch_id = %s")
        params.append(branch_id)
        conds.append("fc.branch_id = %s")
        params.append(branch_id)
    else:
        conds.append("fc.branch_id IS NULL")
    if godown_id:
        conds.append("sm.godown_id = %s")
        params.append(godown_id)
    else:
        conds.append("sm.godown_id IS NULL")
    return " AND ".join(conds), params


# GET /reorder
@router.get("/")
async def get_reorder(
    branch_id: str = "",
    godown_id: str = "",
    user: dict = Depends(require_role("tenant_admin", "tenant_user")),
    db=Depends(get_tenant_db),
):
    """Return bucket1 (priority) and bucket2 (broad scan) for the reorder screen."""
    pub = await get_public_pool()
    tenant_row = await fetchone(pub,
        "SELECT lead_time_days FROM tenants WHERE id = %s", (user["tenantId"],))
    lead_time = int((tenant_row or {}).get("lead_time_days") or 15)

    bg_cond, bg_params = _branch_godown_cond(branch_id, godown_id)

    # Bucket 1: SKUs sold in the last 7 days
    b1_rows = await fetchall(db,
        f"""{_REORDER_SELECT}
            {_REORDER_FROM}
            WHERE {bg_cond}
              AND s.is_active = TRUE
              AND EXISTS (
                SELECT 1 FROM sales sal
                WHERE sal.sku_id = s.id
                  AND sal.sale_date >= NOW() - INTERVAL '7 days'
                  {'AND sal.branch_id = %s' if branch_id else ''}
              )
            ORDER BY fc.woi ASC NULLS LAST""",
        bg_params + ([branch_id] if branch_id else [])
    )
    b1_ids = {str(r["sku_id"]) for r in b1_rows}

    # Bucket 2: all other at-risk SKUs not in Bucket 1
    b2_rows = await fetchall(db,
        f"""{_REORDER_SELECT}
            {_REORDER_FROM}
            WHERE {bg_cond}
              AND s.is_active = TRUE
              AND s.id NOT IN ({','.join(['%s'] * len(b1_ids)) if b1_ids else 'NULL'})
              AND (fc.woi_status IN ('red','amber') OR fc.current_stock <= 0)
            ORDER BY fc.woi ASC NULLS LAST""",
        bg_params + list(b1_ids)
    )

    def _enrich(rows):
        out = []
        for r in rows:
            r = dict(r)
            r["suggested_order_qty"] = _suggested_qty(
                r.get("drr_recommended"), lead_time, r.get("effective_stock"))
            # Determine reorder status
            if r.get("order_status") in ("order_placed", "pending_delivery"):
                r["reorder_status"] = r["order_status"]
            elif float(r.get("effective_stock") or 0) == 0:
                r["reorder_status"] = "out_of_stock"
            else:
                msl = float(r.get("msl") or 0)
                msl_sug = float(r.get("msl_suggested") or 0)
                if msl > 0 and msl_sug > 0 and abs(msl_sug - msl) / msl > 0.20:
                    r["reorder_status"] = "msl_review_recommended"
                else:
                    r["reorder_status"] = "reorder_suggested"
            out.append(r)
        return out

    return {
        "bucket1": _enrich(b1_rows),
        "bucket2": _enrich(b2_rows),
        "summary": {
            "bucket1_count": len(b1_rows),
            "bucket2_count": len(b2_rows),
            "lead_time_days": lead_time,
        },
    }


# ── Reorder Orders ────────────────────────────────────────────────────────────

class OrderCreate(BaseModel):
    sku_id: str
    branch_id: str
    godown_id: Optional[str] = None
    ordered_qty: int
    use_system_lead_time: bool = True
    expected_delivery_dt: Optional[str] = None   # YYYY-MM-DD, only when use_system_lead_time=False
    notes: Optional[str] = None


# POST /reorder/orders  — mark order placed
@router.post("/orders")
async def create_order(
    body: OrderCreate,
    user: dict = Depends(require_role("tenant_admin", "tenant_user")),
    db=Depends(get_tenant_db),
):
    if body.ordered_qty <= 0:
        raise HTTPException(status_code=400, detail="ordered_qty must be > 0")

    # Resolve expected_delivery_dt
    if body.use_system_lead_time:
        pub = await get_public_pool()
        tenant_row = await fetchone(pub,
            "SELECT lead_time_days FROM tenants WHERE id = %s", (user["tenantId"],))
        lead_time = int((tenant_row or {}).get("lead_time_days") or 15)
        delivery_dt = str(_expected_delivery(lead_time))
    else:
        if not body.expected_delivery_dt:
            raise HTTPException(status_code=400,
                detail="expected_delivery_dt is required when use_system_lead_time=False")
        delivery_dt = body.expected_delivery_dt

    new_id = str(uuid.uuid4())
    await execute(db,
        """INSERT INTO skus_reorder_orders
               (sku_id, branch_id, godown_id, ordered_qty, placed_by,
                use_system_lead_time, expected_delivery_dt, status, notes,
                created_at, updated_at)
           VALUES (%s,%s,%s,%s,%s,%s,%s,'order_placed',%s,NOW(),NOW())""",
        (body.sku_id, body.branch_id, body.godown_id, body.ordered_qty,
         user["userId"], body.use_system_lead_time, delivery_dt, body.notes)
    )
    return {"message": "Order placed", "id": new_id, "expected_delivery_dt": delivery_dt}


class OrderUpdate(BaseModel):
    ordered_qty: Optional[int] = None
    expected_delivery_dt: Optional[str] = None
    status: Optional[str] = None   # order_placed | pending_delivery | delivered | cancelled
    notes: Optional[str] = None


# PATCH /reorder/orders/{id}
@router.patch("/orders/{order_id}")
async def update_order(
    order_id: str, body: OrderUpdate,
    user: dict = Depends(require_role("tenant_admin", "tenant_user")),
    db=Depends(get_tenant_db),
):
    updates = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")
    allowed_statuses = {"order_placed", "pending_delivery", "delivered", "cancelled"}
    if "status" in updates and updates["status"] not in allowed_statuses:
        raise HTTPException(status_code=400, detail=f"status must be one of {allowed_statuses}")

    updates["updated_at"] = "NOW()"
    set_parts = []
    params = []
    for k, v in updates.items():
        if v == "NOW()":
            set_parts.append(f"{k} = NOW()")
        else:
            set_parts.append(f"{k} = %s")
            params.append(v)
    params.append(order_id)

    await execute(db,
        f"UPDATE skus_reorder_orders SET {', '.join(set_parts)} WHERE id = %s", params)
    return {"message": "Order updated"}


# GET /reorder/orders  — list open orders
@router.get("/orders")
async def list_orders(
    branch_id: str = "", godown_id: str = "",
    status: str = "",
    page: int = 1, limit: int = 50,
    user: dict = Depends(require_role("tenant_admin", "tenant_user")),
    db=Depends(get_tenant_db),
):
    offset = (page - 1) * limit
    where, params = "WHERE 1=1", []
    if branch_id:
        where += " AND ro.branch_id = %s"; params.append(branch_id)
    if godown_id:
        where += " AND ro.godown_id = %s"; params.append(godown_id)
    if status:
        where += " AND ro.status = %s"; params.append(status)
    else:
        where += " AND ro.status IN ('order_placed','pending_delivery')"

    rows = await fetchall(db,
        f"""SELECT ro.id, ro.sku_id, s.sku_code, s.sku_name,
                   fb.branch_name, g.godown_name,
                   ro.ordered_qty, ro.order_placed_at,
                   ro.expected_delivery_dt, ro.status, ro.notes,
                   u.name AS placed_by_name
            FROM skus_reorder_orders ro
            JOIN skus s ON s.id = ro.sku_id
            JOIN branches fb ON fb.id = ro.branch_id
            LEFT JOIN godowns g ON g.id = ro.godown_id
            LEFT JOIN users u ON u.id = ro.placed_by
            {where}
            ORDER BY ro.order_placed_at DESC LIMIT %s OFFSET %s""",
        params + [limit, offset]
    )
    total_row = await fetchone(db,
        f"SELECT COUNT(*) AS total FROM skus_reorder_orders ro {where}", params)
    return {"data": rows, "total": total_row["total"], "page": page, "limit": limit}


# ── Bulk Export / Upload ───────────────────────────────────────────────────────

_BULK_STATUS_MAP = {
    "ordered":          "order_placed",
    "order placed":     "order_placed",
    "pending":          "pending_delivery",
    "pending delivery": "pending_delivery",
    "delivered":        "delivered",
    "cancelled":        "cancelled",
    "canceled":         "cancelled",
    "ignored":          None,   # None = skip / do nothing
}

_PRIMARY_FILL = PatternFill(fill_type="solid", fgColor="1A3C5E")
_WHITE_FONT   = Font(bold=True, color="FFFFFF")
_CENTER       = Alignment(horizontal="center", vertical="center")


def _build_reorder_excel(rows: list[dict], lead_time: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Smart Reorder"

    headers = [
        ("SKU Code", 16), ("SKU Name", 32), ("Brand", 16),
        ("Stock", 10), ("DRR/day", 10), ("WOI (wks)", 10),
        ("Sugg. Qty", 12), ("Current Status", 20),
        ("Action (Ordered/Pending/Ignored)", 32),
    ]
    ws.append([h[0] for h in headers])
    for i, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = _WHITE_FONT
        cell.fill = _PRIMARY_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 20

    for r in rows:
        status = r.get("order_status") or r.get("reorder_status") or "reorder_suggested"
        ws.append([
            r.get("sku_code"),
            r.get("sku_name"),
            r.get("brand") or "",
            round(float(r.get("effective_stock") or 0), 0),
            round(float(r.get("drr_recommended") or 0), 2),
            round(float(r.get("woi") or 0), 1),
            r.get("suggested_order_qty") or 0,
            status,
            "",   # Action column — user fills this in
        ])

    # Add a data-validation note in the header comment
    ws["I1"].comment = None  # no openpyxl comment needed — label is self-explanatory
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# GET /reorder/export
@router.get("/export")
async def reorder_export(
    branch_id: str = "", godown_id: str = "",
    user: dict = Depends(require_role("tenant_admin", "tenant_user")),
    db=Depends(get_tenant_db),
):
    """Download the current reorder list as Excel for bulk status update."""
    pub = await get_public_pool()
    tenant_row = await fetchone(pub,
        "SELECT lead_time_days FROM tenants WHERE id = %s", (user["tenantId"],))
    lead_time = int((tenant_row or {}).get("lead_time_days") or 15)

    bg_cond, bg_params = _branch_godown_cond(branch_id, godown_id)

    b1_rows = await fetchall(db,
        f"""{_REORDER_SELECT}
            {_REORDER_FROM}
            WHERE {bg_cond} AND s.is_active = TRUE
              AND EXISTS (
                SELECT 1 FROM sales sal WHERE sal.sku_id = s.id
                  AND sal.sale_date >= NOW() - INTERVAL '7 days'
                  {'AND sal.branch_id = %s' if branch_id else ''}
              )
            ORDER BY fc.woi ASC NULLS LAST""",
        bg_params + ([branch_id] if branch_id else [])
    )
    b1_ids = {str(r["sku_id"]) for r in b1_rows}

    b2_rows = await fetchall(db,
        f"""{_REORDER_SELECT}
            {_REORDER_FROM}
            WHERE {bg_cond} AND s.is_active = TRUE
              AND s.id NOT IN ({','.join(['%s'] * len(b1_ids)) if b1_ids else 'NULL'})
              AND (fc.woi_status IN ('red','amber') OR fc.current_stock <= 0)
            ORDER BY fc.woi ASC NULLS LAST""",
        bg_params + list(b1_ids)
    )

    def _enrich(rows):
        out = []
        for r in rows:
            r = dict(r)
            r["suggested_order_qty"] = _suggested_qty(
                r.get("drr_recommended"), lead_time, r.get("effective_stock"))
            if r.get("order_status") in ("order_placed", "pending_delivery"):
                r["reorder_status"] = r["order_status"]
            elif float(r.get("effective_stock") or 0) == 0:
                r["reorder_status"] = "out_of_stock"
            else:
                r["reorder_status"] = "reorder_suggested"
            out.append(r)
        return out

    all_rows = _enrich(b1_rows) + _enrich(b2_rows)
    excel_bytes = _build_reorder_excel(all_rows, lead_time)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=smart_reorder.xlsx"},
    )


# POST /reorder/bulk-upload
@router.post("/bulk-upload")
async def reorder_bulk_upload(
    file: UploadFile = File(...),
    branch_id: str = "",
    user: dict = Depends(require_role("tenant_admin", "tenant_user")),
    db=Depends(get_tenant_db),
):
    """Accept the filled-in reorder Excel; create/update orders based on the Action column."""
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are accepted")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the uploaded file")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Expected columns: SKU Code(0), SKU Name(1), Brand(2), Stock(3),
    #                   DRR(4), WOI(5), Sugg Qty(6), Current Status(7), Action(8)

    pub = await get_public_pool()
    tenant_row = await fetchone(pub,
        "SELECT lead_time_days FROM tenants WHERE id = %s", (user["tenantId"],))
    lead_time = int((tenant_row or {}).get("lead_time_days") or 15)
    delivery_dt = str(_expected_delivery(lead_time))

    success, skipped, errors = 0, 0, []

    for row_num, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            continue

        sku_code = str(row[0]).strip() if row[0] else ""
        action_raw = str(row[8]).strip().lower() if len(row) > 8 and row[8] else ""

        if not sku_code:
            continue
        if not action_raw or action_raw in ("", "none", "-", "—"):
            skipped += 1
            continue

        new_status = _BULK_STATUS_MAP.get(action_raw)
        if new_status is False or (new_status is None and action_raw not in _BULK_STATUS_MAP):
            errors.append(f"Row {row_num}: unknown action '{row[8]}'")
            continue

        if new_status is None:  # "ignored" — skip
            skipped += 1
            continue

        # Lookup SKU
        sku_row = await fetchone(db,
            "SELECT id FROM skus WHERE sku_code = %s AND is_active = TRUE", (sku_code,))
        if not sku_row:
            errors.append(f"Row {row_num}: SKU '{sku_code}' not found")
            continue

        sku_id = str(sku_row["id"])

        # Check for existing open order
        existing = await fetchone(db,
            """SELECT id FROM skus_reorder_orders
               WHERE sku_id = %s AND status NOT IN ('delivered','cancelled')
               ORDER BY created_at DESC LIMIT 1""",
            (sku_id,)
        )

        if existing:
            await execute(db,
                "UPDATE skus_reorder_orders SET status = %s, updated_at = NOW() WHERE id = %s",
                (new_status, str(existing["id"]))
            )
        else:
            # Create a new order with suggested_order_qty = 1 as minimum placeholder
            sugg_qty = int(row[6]) if len(row) > 6 and row[6] else 1
            await execute(db,
                """INSERT INTO skus_reorder_orders
                       (sku_id, branch_id, ordered_qty, placed_by,
                        use_system_lead_time, expected_delivery_dt, status, created_at, updated_at)
                   VALUES (%s, %s, %s, %s, TRUE, %s, %s, NOW(), NOW())""",
                (sku_id, branch_id or None, max(1, sugg_qty),
                 user["userId"], delivery_dt, new_status)
            )
        success += 1

    return {
        "message": f"Bulk update complete: {success} updated, {skipped} skipped, {len(errors)} errors",
        "success": success,
        "skipped": skipped,
        "errors": errors,
    }
